__author__ = 'stephenmidgley'
from selenium import webdriver
import time
import datetime
import csv
import re
from openpyxl import load_workbook
from urlparse import urlparse
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait # available since 2.4.0
from selenium.webdriver.support import expected_conditions as EC # available since 2.26.0
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time






wb = load_workbook("AllJobs.xlsx")
ws = wb.active
end = True
PreviousList = []






#CWJobs Bit
def getJobsFromCWJobsURL(url,Score):
    Jobs = []
    driver.get(url)
    results = driver.find_elements_by_class_name("job-result")
    for result in results:
        all_children_by_xpath = result.find_elements_by_xpath(".//*")
        job = {}
        for child in all_children_by_xpath:
            if child is all_children_by_xpath[0]:
                job['title']=child.text
            if child.get_attribute("class") == 'jd':
                job['description'] = child.text
            href = child.get_attribute("href")
            if href is not None:
                url  = urlparse(href)
                querystring = url.query
                if len(querystring)>5:

                    queryArray = querystring.split("&")
                    for argument in queryArray:
                        if argument.split("=")[0] == 'JobId':
                            job['JobId']=argument.split("=")[1]
                            job['Score']=Score
                            job['url']= href
                        if argument.split("=")[0] == 'Keywords':
                            job['Keywords'] =argument.split("=")[1]



                        #job['Description'] = description.getText()
                        #print description.getText()


        Jobs.append(job)


    return Jobs



def getCWJObsURLs(keyword, location):
        # go to the google home page
        driver.get("http://www.cwjobs.co.uk")
        # the page is ajaxy so the title is originally this:
        SearchTerm = keyword
        Location = location
        # find the element that's name attribute is q (the google search box)
        inputElement = driver.find_element_by_id("keywords")
        # type in the search
        inputElement.clear()
        inputElement.send_keys(SearchTerm)
        inputElement = driver.find_element_by_id("location")
        inputElement.clear()
        inputElement.send_keys(Location)

        # submit the form (although google automatically searches now without submitting)
        searchbutton = driver.find_element_by_id('search-button')

        searchbutton.click()
        time.sleep(5)


        contractfield = driver.find_elements_by_xpath("//*[contains(text(), 'Contract')]")
        nocontracts = False
        try:
            elementclass =  contractfield[0].get_attribute("class")
        except:
            elementclass="unavailable"

        if "unavailable" in elementclass:
           # print "aww shucks no contracts for " + SearchTerm
            nocontracts = True
        else:
          #  print "Found some potential contacts in " + SearchTerm
            href = contractfield[0].get_attribute("href")
            if href is not None:
                contractfield[0].click()
                time.sleep(5)
                hrefstopages = [href]
                ListlinkerHref = driver.find_elements_by_xpath("//*[@href]")
                for hrefelwment in ListlinkerHref:
                    if "PageNum"  in hrefelwment .get_attribute("href"):
                        hrefstopages.append( hrefelwment.get_attribute("href"))
                  #Need to check for duplicates in array,
                hrefstopages = list(set(hrefstopages))
                return hrefstopages

                #Now just need to create a simple array look for looping through each of the pages....9

            else:
                return []


def getJobsFromReed (searchTerm,location, Score):

    url = "http://www.reed.co.uk/"
    jobs = []
    driver.get(url)
    time.sleep(1)

    try:
        inputElement = driver.find_element_by_id("main-keywords")
    except:
        try:
            inputElement = driver.find_element_by_id("keywords")
        except:
            print "Trouble finding contract" + searchTerm
            return []
    # type in the search
    inputElement.clear()
    inputElement.send_keys(searchTerm)
    try:
        inputElement = driver.find_element_by_id("main-location")
    except:
        try:
            inputElement = driver.find_element_by_id("location")
        except:
            print "Trouble finding contract" + searchTerm
            return []
    inputElement.clear()
    inputElement.send_keys(location)

    try:
        searchbutton = driver.find_element_by_id('homepageSearchButton')
    except:
        try:
            searchbutton = driver.find_element_by_id('SearchButton')
        except:
         #   print "Trouble finding contract" + searchTerm
            return []


    try:
        searchbutton.click()
    except:
        #print "Trouble finding contract" + searchTerm
        return []

    try:
        contract = driver.find_element_by_link_text('Contract')
        link = contract.get_attribute('href')
        if "contract=True" in link:
            driver.get(link)
            try:
                element = WebDriverWait(driver, 5).until(
                    EC.presence_of_element_located((By.CLASS_NAME, "job-result"))
                )
            except:
                print "Reed Timeout Issue"
                pass

         #   print "Found contract(s) for " + searchTerm

            results = driver.find_elements_by_class_name("job-result")
            for result in results:
                driver.save_screenshot('screenie.png')
                all_children_by_xpath = result.find_elements_by_xpath(".//*")
                job = {}
                job['JobId']=result.get_attribute('id')
                job['Score']=Score
                job['Keywords'] = searchTerm
                for child in all_children_by_xpath:

                    if child.get_attribute("class") == 'description':
                        job['description'] = child.text

                    if child.get_attribute("class") == 'title':
                        job['title'] = child.text

                        titleChildren = child.find_elements_by_xpath(".//*")
                        for titlechild in titleChildren:
                            if titlechild.get_attribute("href") is not None:
                                job['url'] = titlechild.get_attribute("href")
                jobs.append(job)




        else:
            pass
        #    print "No contracts for job" + searchTerm
    except ValueError:
        pass
       # print ValueError
       # print "No contracts for job" + searchTerm
    #print jobs
    return jobs


def getJobsfromContractorUK(searchTerm,Location, Score):
    url = "http://www.contractoruk.com/contract_search_wizard/"
    jobs = []
    driver.get(url)
    time.sleep(1)

    try:
        inputElement = driver.find_element_by_id("edit-keywords")
    except:
        try:
            inputElement = driver.find_element_by_id("keywords")
        except:
           # print "Trouble finding contract" + searchTerm
            return []
    inputElement.clear()
    inputElement.send_keys(searchTerm)
    try:
        inputElement = driver.find_element_by_id("edit-location")
    except:
        try:
            inputElement = driver.find_element_by_id("location")
        except:
          #  print "Trouble finding contract" + searchTerm
            return []
    inputElement.clear()
    inputElement.send_keys(Location)

    try:
        searchbutton = driver.find_element_by_id('edit-submit')
    except:
        try:
            searchbutton = driver.find_element_by_id('edit-submit')
        except:
            #print "Trouble finding contract" + searchTerm
            return []


    try:
        searchbutton.click()
    except:
        #print "Trouble finding contract" + searchTerm
        return []



    while True:
        try:
            element = WebDriverWait(driver, 20).until(
                EC.presence_of_element_located((By.CLASS_NAME, "csw-result-premium"))
            )
        except:
            #print "No contracts found in IT contract finder " +searchTerm
            return []
        results = driver.find_elements_by_class_name("csw-result-premium")
        for result in results:
                    all_children_by_xpath = result.find_elements_by_xpath(".//*")
                    #print all_children_by_xpath
                    job = {}
                    url = all_children_by_xpath[1].get_attribute("href")

                    if re.search('[/]\d{5,}',url):

                        match = re.search('[/]\d{5,}',url)
                     #   print "got ID" + match.group(0)

                        job['JobId'] = match.group(0)
                    else:
                      #  print "no Match just use the old format"
                        job['JobId']=all_children_by_xpath[4].text



                    job['Score']=Score
                    job['Keywords'] = searchTerm
                    job['description'] = all_children_by_xpath[4].text
                    job['title'] = all_children_by_xpath[0].text
                    job['url'] = all_children_by_xpath[1].get_attribute("href")
                    jobs.append(job)
        try:
            element = driver.find_element_by_link_text("Next >>")
            element.click()

        except:
            return jobs


    #<h1> Is the title, next bit of text is the description






# Create a new instance of the Firefox driver

driver = webdriver.Firefox() # or add to your PATH
driver.set_window_size(1024, 768) # optional
#driver = webdriver.Firefox()
driver.implicitly_wait(10)

JobList = []
#SearchTerms and Score....
#SearchTerms = ['eLearning','Technical Trainer','Technical Project Manager', 'Rest API', 'RESTAPI','GX','elearning','e-learning','Web tools','webtools', 'HTML 5 Telecoms','SIP ','HTML5 Telecoms','angularjs','Telecommunications']
SearchTerms = [
                ['Technical Trainer',10],
                 ['REST-API Trainer',10],
                 ['REST API Trainer',10],
                 ['API Trainer',10],
                 ['Telecoms Trainer',10],
                 ['Technical Lead Telecoms',10],
                 ['Technical Lead eLearning',10],
                 ['Technical Lead e-Learning',10],
                 ['Technical Lead Telecommunications',10],
                 ['Pre-Sales Telecoms Consultant',10],
                 ['Pre-Sales Telecommunications Consultant',10],
                 ['Pre-Sales Consultant',9],
                 ['Telephony Consultant',9],
                 ['Technical Consultant',6],
                 ['Pre-Sales',8],
                 ['Technical Project Manager Telecoms',10],
                 ['Technical Project Manager Telecommunications',10],
                 ['Technical Project Manager HTML5',7],
                 ['Technical Project Manager',5],
                ['e-Learning',7],
              ['eLearning',7],

                 ['Webapps',6],
                 ['Webtools',6],
                 ['Instructional Designer Telecoms',10],
                ['Instructional Designer Telecommunications',10],
                 ['Instructional Designer',5],
                 ['Business Analyst Telecoms',10],
                 ['Business Analyst Telecommunications',10],
                 ['Business Analyst Satcoms',10],
                 ['Business Analyst',5],
                 ['Trainer SIP',10],
                 ['Python Telecoms',10],
                 ['Python Telecommunications',10],
                 ['Python Trainer',9],
                 ['GX API',10],
                 ['GX',5],
                 ['HTML Telecoms',3],
                 ['HTML5 Telecoms',3],
                 ['Telephony',3],
                 ['Devops',6],
                 ['Devops Networking',15],
                 ['Devops Networking Javascript',20],
                 ['Devops Telecoms',10],
                 ['Devops Telecommunications',10],
                 ['Devops Trainer',15]
               ]



y=0.0
Terms = len(SearchTerms)
start_time = time.time()
for searchterm in SearchTerms:
    print "Searching for " +searchterm[0]
    print str(y/Terms * 100.0) + "% Done"

    if y> 0:
        print ((1/(y/Terms)) * (time.time() - start_time))/60, "Minutes Left"
    urls = getCWJObsURLs(searchterm[0], 'SE192UP')
    try:
         for url in urls:
            JobList.extend(getJobsFromCWJobsURL(url,searchterm[1]))
    except:
         pass

    JobList.extend(getJobsFromReed(searchterm[0],'SE192UP',searchterm[1]))
    JobList.extend(getJobsfromContractorUK(searchterm[0],'SE192UP',searchterm[1]))
    y=y+1


with open('TodaysJob.csv', 'w') as csvfile:
    fieldnames = ['title','description', 'url','JobId','Keywords','Score']
    writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
    writer.writeheader()
    for job in JobList:
        job['title']= job['title'].encode('ascii', 'ignore')
        job['Score']= str(job['Score']).encode('ascii', 'ignore')
        job['description']= job['description'].encode('ascii', 'ignore')
        job['url']= job['url'].encode('ascii', 'ignore')
        job['JobId']= job['JobId'].encode('ascii', 'ignore')
        job['Keywords']= job['Keywords'].encode('ascii', 'ignore')
        try:
            writer.writerow(job)
        except:
            pass



#This is a test

#Load the existing excel sheet
wb = load_workbook("AllJobs.xlsx")
ws = wb.active
PreviousList = []
x=0
while True:
    x = x + 1
    cell = ws.cell('f'+ str(x))
    if cell.value != None:
        print "Compiling previous list with" + str(cell.value)
        PreviousList.append( cell.value)
    else:
        break

for job in JobList:
    if job['JobId'] not in PreviousList:
        #Adding to excel sheet
        print "Adding job" + job['JobId']
        ws['e'+ str(x)] = job['JobId']
        ws['a'+ str(x)] =job['title']
        ws['b'+ str(x)]=job['description']
        ws['e'+ str(x)]=job['url']
        ws['f'+ str(x)]=job['JobId']
        ws['g'+ str(x)]=job['Keywords']
        ws['h'+ str(x)]=job['Score']
        ws['c'+ str(x)] = str(datetime.date.today())
        x = x +1
    else:
        print 'Job already found ignoring'
    #Add to excel sheet
wb.save("AllJobs.xlsx")
#Check if the
